from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query, Body
from pydantic import ValidationError
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session
from sqlalchemy import func, extract
from datetime import datetime, date, timedelta
from uuid import UUID
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
import io, re

from app.models.project import Project
from app.models.site import Site

from app.db.session import SessionLocal
from app.models.worker import Worker
from app.models import project
from app.models.attendance import AttendanceRecord
from app.models.site import Site
from app.models.payroll import PayrollEntry
from app.schemas.worker import WorkerCreate, WorkerResponse, WorkerUpdate, ActionResponse, ForceDeleteRequest, ArchiveRequest, WorkerBulkRequest
from app.core.dependencies import require_admin
from app.services.audit_service import log_action
from app.services.r2 import generate_presigned_download_url, upload_file_to_r2, s3, R2_BUCKET
from app.services.image_service import format_site_folder
from app.services.worker_photo_service import resolve_worker_profile_photo_key

router = APIRouter(prefix="/workers", tags=["Workers"])


# -----------------------------
# DB Dependency
# -----------------------------
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


# -----------------------------
# Audit helper — extracts name/role/id from JWT payload dict
# -----------------------------
def _audit_user(user: dict) -> dict:
    """Return kwargs for log_action from a JWT payload dict."""
    return {
        "performed_by": UUID(user["sub"]) if user.get("sub") else None,
        "performed_by_name": user.get("name") or user.get("sub", "Unknown"),
        "performed_by_role": user.get("role", "admin"),
    }


# --------------------------------------------------
# CREATE WORKER
# --------------------------------------------------

@router.post("/", response_model=WorkerResponse)
def create_worker(data: WorkerCreate, db: Session = Depends(get_db), current_user: dict = Depends(require_admin)):

    payload = data.dict()

    # Generate Custom Worker ID
    last3_mobile = payload["mobile"][-3:]
    last3_id = payload["id_number"][-3:]
    worker_id = f"E{last3_mobile}{last3_id}"

    # Prevent Worker ID conflict
    existing = db.query(Worker).filter(Worker.id == worker_id).first()
    if existing:
        raise HTTPException(
            status_code=400,
            detail="Worker ID conflict. Mobile + ID combination must be unique."
        )

    # Check duplicate mobile
    mobile_exists = db.query(Worker).filter(Worker.mobile == payload["mobile"]).first()
    if mobile_exists:
        raise HTTPException(
            status_code=400,
            detail="This mobile number is already registered."
        )

    # Check duplicate ID number
    id_exists = db.query(Worker).filter(Worker.id_number == payload["id_number"]).first()
    if id_exists:
        raise HTTPException(
            status_code=400,
            detail="This ID number is already registered."
        )

    try:

        worker = Worker(
            id=worker_id,
            joining_date=date.today(),
            **payload
        )

        db.add(worker)
        db.commit()
        db.refresh(worker)

    except IntegrityError:

        db.rollback()

        raise HTTPException(
            status_code=400,
            detail="Database error while creating worker."
        )

    log_action(
        db=db,
        action="create",
        entity_type="worker",
        entity_id=worker.id,
        details=f"Worker {worker.full_name} created",
        **_audit_user(current_user),
    )

    return worker


# --------------------------------------------------
# WORKER TEMPLATE ENDPOINTS
# --------------------------------------------------

MAX_BULK_ROWS = 100


def sanitize_name(name: str):
    """Convert project name into Excel-safe named range"""
    name = re.sub(r"\W+", "_", name)
    if name[0].isdigit():
        name = "_" + name
    return name


@router.get("/template-excel")
def download_worker_template(db: Session = Depends(get_db)):

    projects = (
        db.query(Project)
        .filter(Project.is_deleted == False, Project.status == "active")
        .order_by(Project.name)
        .all()
    )

    sites = (
        db.query(Site)
        .filter(Site.is_deleted == False, Site.status == "active")
        .order_by(Site.name)
        .all()
    )

    wb = Workbook()

    # -----------------------
    # workers sheet
    # -----------------------

    ws = wb.active
    ws.title = "workers"

    headers = [
        "full_name",
        "mobile",
        "id_number",
        "project",
        "site",
        "role",
        "type",
        "monthly_salary",
        "daily_rate",
        "daily_working_hours",
        "hourly_rate",
        "ot_multiplier",
    ]

    ws.append(headers)

    for _ in range(MAX_BULK_ROWS):
        ws.append([""] * len(headers))

    # -----------------------
    # projects sheet
    # -----------------------

    ws_projects = wb.create_sheet("projects")
    ws_projects.append(["project"])

    for p in projects:
        ws_projects.append([p.name])

    # -----------------------
    # roles sheet
    # -----------------------

    ws_roles = wb.create_sheet("roles")

    roles = ["Electrician", "Fitter", "Mason", "Labourer", "Foreman", "Site Manager", "Engineer", "Technician", "Helper", "Site In-Charge", "Supervisor", "Other"]

    ws_roles.append(["role"])

    for r in roles:
        ws_roles.append([r])

    # -----------------------
    # sites sheet
    # -----------------------

    ws_sites = wb.create_sheet("sites")

    project_map = {p.id: p for p in projects}

    project_sites = {}

    for p in projects:
        project_sites[p.name] = []

    for s in sites:
        project = project_map.get(s.project_id)

        if not project:
            continue
        
        project_name = project.name
        project_sites[project_name].append(s.name)

    col = 1

    for project_name, site_list in project_sites.items():

        safe_name = sanitize_name(project_name)

        ws_sites.cell(row=1, column=col).value = safe_name

        for i, site in enumerate(site_list, start=2):
            ws_sites.cell(row=i, column=col).value = site

        start = f"${ws_sites.cell(row=2, column=col).column_letter}$2"
        end = f"${ws_sites.cell(row=len(site_list)+1, column=col).column_letter}${len(site_list)+1}"

        wb.defined_names.add(
            DefinedName(
                name=safe_name,
                attr_text=f"sites!{start}:{end}",
            )
        )

        col += 1

    # -----------------------
    # dropdown validations
    # -----------------------

    project_validation = DataValidation(
        type="list",
        formula1=f"=projects!$A$2:$A${len(projects)+1}",
        allow_blank=False,
    )

    ws.add_data_validation(project_validation)
    project_validation.add(f"D2:D{MAX_BULK_ROWS+1}")

    role_validation = DataValidation(
        type="list",
        formula1=f"=roles!$A$2:$A${len(roles)+1}",
    )

    ws.add_data_validation(role_validation)
    role_validation.add(f"F2:F{MAX_BULK_ROWS+1}")

    type_validation = DataValidation(
        type="list",
        formula1='"contract,permanent"',
    )

    ws.add_data_validation(type_validation)
    type_validation.add(f"G2:G{MAX_BULK_ROWS+1}")

    # -----------------------
    # formulas
    # -----------------------

    for row in range(2, MAX_BULK_ROWS + 2):

        # site dependent dropdown
        dv_site = DataValidation(
            type="list",
            formula1=f'=INDIRECT(SUBSTITUTE(D{row}," ","_"))'
        )
        ws.add_data_validation(dv_site)
        dv_site.add(f"E{row}")

        # auto fill first site
        ws[f"E{row}"] = f'=IF(D{row}="","",INDEX(INDIRECT(SUBSTITUTE(D{row}," ","_")),1))'

        # daily rate calculation
        ws[f"I{row}"] = f'=IF(G{row}="permanent",IF(H{row}<>"",ROUND(H{row}/30,2),""),"")'

        # hourly rate
        ws[f"K{row}"] = f'=IF(AND(I{row}<>"",J{row}<>""),ROUND(I{row}/J{row},2),"")'


    # -----------------------
    # export
    # -----------------------

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=workers_template.xlsx"
        },
    )


# --------------------------------------------------
# DELETE WORKER
# --------------------------------------------------
@router.delete("/{worker_id}")
def delete_worker(worker_id: str, db: Session = Depends(get_db), current_user: dict = Depends(require_admin)):
    worker = db.query(Worker).filter(Worker.id == worker_id).first()

    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")

    db.delete(worker)
    db.commit()

    log_action(
        db=db,
        action="delete",
        entity_type="worker",
        entity_id=worker.id,
        details=f"Worker {worker.full_name} deleted",
        **_audit_user(current_user),
    )

    return {"message": "Worker deleted successfully"}


# --------------------------------------------------
# LIST WORKERS
# --------------------------------------------------


@router.get("/", response_model=list[WorkerResponse], dependencies=[Depends(require_admin)])
def list_workers(
    include_deleted: bool = False,
    project_id: str | None = None,
    site_id: str | None = None,
    db: Session = Depends(get_db)
):
    query = db.query(Worker)

    if not include_deleted:
        query = query.filter(Worker.is_deleted == False)

    if project_id:
        query = query.filter(Worker.project_id == project_id)

    if site_id:
        query = query.filter(Worker.site_id == site_id)

    workers = query.all()

    for w in workers:
        if w.photo_url:
            w.photo_signed_url = generate_presigned_download_url(w.photo_url)
        else:
            w.photo_signed_url = None

    return workers


# --------------------------------------------------
# GET SINGLE WORKER
# --------------------------------------------------
@router.get("/{worker_id}", response_model=WorkerResponse, dependencies=[Depends(require_admin)])
def get_worker(worker_id: str, db: Session = Depends(get_db)):
    worker = db.query(Worker).filter(Worker.id == worker_id).first()

    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")

    return worker


# --------------------------------------------------
# WORKER ATTENDANCE SUMMARY
# --------------------------------------------------
@router.get("/{worker_id}/attendance-summary", dependencies=[Depends(require_admin)])
def worker_attendance_summary(worker_id: str, db: Session = Depends(get_db)):

    present_days = db.query(AttendanceRecord)\
        .filter(
            AttendanceRecord.worker_id == worker_id,
            AttendanceRecord.check_in_time.isnot(None)
        ).count()

    total_hours = db.query(
        func.sum(AttendanceRecord.total_hours)
    ).filter(
        AttendanceRecord.worker_id == worker_id
    ).scalar() or 0

    absent_days = 0  # Can be implemented later

    return {
        "present_days": present_days,
        "absent_days": absent_days,
        "total_hours": total_hours
    }


# --------------------------------------------------
# WORKER ATTENDANCE INSIGHT  (for profile tab)
# --------------------------------------------------
@router.get("/{worker_id}/attendance-insight", dependencies=[Depends(require_admin)])
def worker_attendance_insight(
    worker_id: str,
    year: int | None = None,
    month: int | None = None,
    db: Session = Depends(get_db),
):
    """
    Returns attendance analytics for the worker profile Attendance tab.
    Optional ?year=&month= to drive the calendar heatmap (defaults to current month).
    """
    today = date.today()
    cal_year  = year  or today.year
    cal_month = month or today.month

    records = (
        db.query(AttendanceRecord)
        .filter(AttendanceRecord.worker_id == worker_id)
        .order_by(AttendanceRecord.date.desc())
        .all()
    )

    # ── Stat Pills ────────────────────────────────────────────────────────────
    checked_in = [r for r in records if r.check_in_time]
    total_present = len(checked_in)
    total_hours   = sum(r.total_hours or 0 for r in records)

    # Exclude today's in-progress attendance (checked-in but not checked out) from Avg/Day denominator.
    checked_out_days = len([r for r in checked_in if r.check_out_time])
    avg_daily_hrs = round(total_hours / checked_out_days, 1) if checked_out_days else 0

    # Avg check-in / check-out times (minutes since midnight)
    def avg_time(times):
        valid = [t.hour * 60 + t.minute for t in times if t]
        if not valid:
            return None
        avg_min = int(sum(valid) / len(valid))
        return f"{avg_min // 60:02d}:{avg_min % 60:02d}"

    avg_checkin  = avg_time([r.check_in_time  for r in checked_in])
    avg_checkout = avg_time([r.check_out_time for r in checked_in if r.check_out_time])

    # Streaks
    present_dates = {r.date for r in records if r.check_in_time}
    streak = 0
    check = today
    while check in present_dates:
        streak += 1
        check -= timedelta(days=1)

    sorted_dates = sorted(present_dates)
    longest = cur = 0
    prev = None
    for d in sorted_dates:
        if prev and (d - prev).days == 1:
            cur += 1
        else:
            cur = 1
        longest = max(longest, cur)
        prev = d

    # ── Monthly Trend (last 6 months) ─────────────────────────────────────────
    monthly: dict[str, float] = {}
    for i in range(5, -1, -1):
        y = today.year
        m = today.month - i
        while m <= 0:
            m += 12
            y -= 1
        monthly[date(y, m, 1).strftime("%b %Y")] = 0.0

    for r in records:
        if not r.check_in_time:
            continue
        label = r.date.strftime("%b %Y")
        if label in monthly:
            monthly[label] += r.total_hours or 0

    monthly_trend = [{"month": k, "hours": round(v, 1)} for k, v in monthly.items()]

    # ── Heatmap — requested month (full month, future days = "future") ─────────
    first_day = date(cal_year, cal_month, 1)
    if cal_month == 12:
        last_day = date(cal_year + 1, 1, 1) - timedelta(days=1)
    else:
        last_day = date(cal_year, cal_month + 1, 1) - timedelta(days=1)

    month_map: dict[date, str] = {}
    for r in records:
        if first_day <= r.date <= last_day and r.check_in_time:
            month_map[r.date] = "late" if r.is_late else "present"

    heatmap = []
    d = first_day
    while d <= last_day:
        if d > today:
            status = "future"
        else:
            status = month_map.get(d, "absent")
        heatmap.append({
            "date":    d.isoformat(),
            "day":     d.day,
            "weekday": d.strftime("%a"),
            "status":  status,
        })
        d += timedelta(days=1)

    return {
        "stat_pills": {
            "total_present":  total_present,
            "total_hours":    round(total_hours, 1),
            "avg_daily_hrs":  avg_daily_hrs,
            "avg_checkin":    avg_checkin,
            "avg_checkout":   avg_checkout,
            "current_streak": streak,
            "longest_streak": longest,
        },
        "monthly_trend": monthly_trend,
        "heatmap": heatmap,
        "heatmap_label": first_day.strftime("%B %Y"),
    }


# --------------------------------------------------
# PATCH / UPDATE WORKER
# --------------------------------------------------

@router.patch("/{worker_id}", response_model=WorkerResponse)
def patch_worker(
    worker_id: str,
    data: WorkerUpdate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_admin),
):
    worker = db.query(Worker).filter(Worker.id == worker_id).first()

    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")

    update_data = data.dict(exclude_unset=True)
    changed_fields = ", ".join(update_data.keys())

    for key, value in update_data.items():
        setattr(worker, key, value)

    worker.updated_at = datetime.utcnow()

    db.commit()
    db.refresh(worker)

    log_action(
        db=db,
        action="update",
        entity_type="worker",
        entity_id=worker.id,
        details=f"Worker {worker.full_name} updated. Fields changed: {changed_fields}",
        **_audit_user(current_user),
    )

    return worker


@router.patch("/{worker_id}/archive",response_model=ActionResponse)
def archive_worker(worker_id: str, payload: ArchiveRequest, db: Session = Depends(get_db), current_user: dict = Depends(require_admin)):
    worker = db.query(Worker).filter(Worker.id == worker_id).first()

    if not worker:
        raise HTTPException(404, "Worker not found")

    if worker.is_deleted:
        raise HTTPException(400, "Worker already archived")

    worker.is_deleted = True
    worker.deleted_at = datetime.utcnow()
    worker.status = "inactive"

    db.commit()

    log_action(
        db=db,
        action="archive",
        entity_type="worker",
        entity_id=worker.id,
        details=payload.reason,
        **_audit_user(current_user),
    )

    return {"message": "Worker archived successfully"}



@router.patch("/{worker_id}/restore",response_model=ActionResponse)
def restore_worker(worker_id: str, db: Session = Depends(get_db), current_user: dict = Depends(require_admin)):

    worker = db.query(Worker).filter(Worker.id == worker_id).first()

    if not worker:
        raise HTTPException(404, "Worker not found")

    if not worker.is_deleted:
        raise HTTPException(400, "Worker is not archived")

    worker.is_deleted = False
    worker.deleted_at = None
    worker.deleted_by = None
    worker.status = "active"

    db.commit()

    log_action(
        db=db,
        action="restore",
        entity_type="worker",
        entity_id=worker.id,
        details=f"Worker {worker.full_name} restored",
        **_audit_user(current_user),
    )

    return {"message": "Worker restored successfully"}



@router.delete("/{worker_id}/force-delete",response_model=ActionResponse)
def force_delete_worker(worker_id: str, payload: ForceDeleteRequest, db: Session = Depends(get_db), current_user: dict = Depends(require_admin)):
    worker = db.query(Worker).filter(Worker.id == worker_id).first()

    if not worker:
        raise HTTPException(404, "Worker not found")

    if payload.confirmation != worker.full_name:
        raise HTTPException(400, "Confirmation text mismatch")

    # Delete attendance linked to worker
    db.query(AttendanceRecord).filter(AttendanceRecord.worker_id == worker_id).delete()

    # Delete worker
    db.delete(worker)
    db.commit()

    log_action(
        db=db,
        action="force_delete",
        entity_type="worker",
        entity_id=worker.id,
        details=payload.reason,
        **_audit_user(current_user),
    )

    return {"message": "Worker permanently deleted"}


# --------------------------------------------------
# WORKER PHOTO
# --------------------------------------------------
@router.get("/{worker_id}/photo")
def get_worker_photo(worker_id: str, db: Session = Depends(get_db)):
    worker = db.query(Worker).filter(Worker.id == worker_id).first()

    if not worker:
        raise HTTPException(404, "Worker not found")

    photo_key = resolve_worker_profile_photo_key(worker, db)
    if not photo_key:
        raise HTTPException(404, "Photo not found")

    url = generate_presigned_download_url(photo_key)

    return {"url": url}


# --------------------------------------------------
# WORKER ASSETS  (R2 folder: {site_folder}/{worker_id}/Assets/)
# --------------------------------------------------

def _worker_asset_prefix(worker: Worker, db: Session) -> str:
    """Return the R2 prefix for this worker's Assets folder."""
    site = db.query(Site).filter(Site.id == worker.site_id).first()
    if not site:
        raise HTTPException(404, "Worker has no site assigned")
    site_folder = format_site_folder(site.name)
    return f"{site_folder}/{worker.id}/Assets/"


@router.get("/{worker_id}/assets", dependencies=[Depends(require_admin)])
def list_worker_assets(worker_id: str, db: Session = Depends(get_db)):
    """List all files in the worker's Assets folder on R2."""
    worker = db.query(Worker).filter(Worker.id == worker_id).first()
    if not worker:
        raise HTTPException(404, "Worker not found")

    prefix = _worker_asset_prefix(worker, db)

    response = s3.list_objects_v2(Bucket=R2_BUCKET, Prefix=prefix)

    files = []
    for obj in response.get("Contents", []):
        key = obj["Key"]
        if key == prefix or key.endswith("/"):
            continue
        files.append({
            "key": key,
            "name": key.split("/")[-1],
            "size": obj["Size"],
            "last_modified": obj["LastModified"].isoformat(),
            "download_url": generate_presigned_download_url(key),
        })

    return {"prefix": prefix, "files": files}


@router.post("/{worker_id}/assets", dependencies=[Depends(require_admin)])
async def upload_worker_asset(
    worker_id: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """Upload a file into the worker's Assets folder on R2."""
    worker = db.query(Worker).filter(Worker.id == worker_id).first()
    if not worker:
        raise HTTPException(404, "Worker not found")

    prefix = _worker_asset_prefix(worker, db)
    object_key = f"{prefix}{file.filename}"

    file_bytes = await file.read()
    upload_file_to_r2(
        file_bytes=file_bytes,
        object_key=object_key,
        content_type=file.content_type or "application/octet-stream",
    )

    return {
        "message": "File uploaded successfully",
        "key": object_key,
        "name": file.filename,
    }


@router.delete("/{worker_id}/assets", dependencies=[Depends(require_admin)])
def delete_worker_asset(worker_id: str, key: str, db: Session = Depends(get_db)):
    """Delete a specific file from the worker's Assets folder on R2."""
    worker = db.query(Worker).filter(Worker.id == worker_id).first()
    if not worker:
        raise HTTPException(404, "Worker not found")

    # Safety: ensure the key actually belongs to this worker
    prefix = _worker_asset_prefix(worker, db)
    if not key.startswith(prefix.rstrip("/")):
        raise HTTPException(403, "Key does not belong to this worker's assets")

    try:
        s3.delete_object(Bucket=R2_BUCKET, Key=key)
    except Exception:
        raise HTTPException(500, "Delete failed")

    return {"message": "File deleted successfully"}


# ──────────────────────────────────────────────────────────────────
#  PAYROLL ENDPOINTS
# ──────────────────────────────────────────────────────────────────

@router.get("/{worker_id}/payroll", dependencies=[Depends(require_admin)])
def get_payroll(
    worker_id: str,
    year:  int = Query(default=None),
    month: int = Query(default=None),
    db: Session = Depends(get_db),
):
    """
    Returns payroll summary for a given month:
    - gross earnings (from attendance × rate)
    - advance/deduction/bonus entries
    - net payable
    - payment history (last 6 months)
    - earnings history (last 6 months) for charts
    """
    from calendar import month_name as _month_name
    import math

    worker = db.query(Worker).filter(Worker.id == worker_id).first()
    if not worker:
        raise HTTPException(404, "Worker not found")

    today = date.today()
    y = year or today.year
    m = month or today.month
    month_label = f"{_month_name[m]} {y}"

    # ── Attendance for the month (finalized only) ───────────────────────────
    records = (
        db.query(AttendanceRecord)
        .filter(
            AttendanceRecord.worker_id == worker_id,
            extract("year", AttendanceRecord.date) == y,
            extract("month", AttendanceRecord.date) == m,
            AttendanceRecord.status == "checked_out",
        )
        .all()
    )

    present_days = len(records)
    total_hours = round(sum(r.total_hours or 0 for r in records), 2)

    # ── Gross calculation ───────────────────────────────────────────────────
    if worker.type == "contract":
        rate = worker.hourly_rate or 0
        rate_label = f"SAR {rate}/hr"
        # Business rule: floor integer hours at payroll level as well
        gross = round(math.floor(total_hours or 0) * rate, 2)
    else:  # permanent
        rate = worker.daily_rate or 0
        rate_label = f"SAR {rate}/day"
        gross = round(present_days * rate, 2)

    # ── Entries for selected month ──────────────────────────────────────────
    month_entries = (
        db.query(PayrollEntry)
        .filter(
            PayrollEntry.worker_id == worker_id,
            PayrollEntry.year == y,
            PayrollEntry.month == m,
        )
        .order_by(PayrollEntry.date.desc())
        .all()
    )

    total_advances = sum(e.amount for e in month_entries if e.entry_type == "advance")
    total_deductions = sum(e.amount for e in month_entries if e.entry_type == "deduction")
    total_bonuses = sum(e.amount for e in month_entries if e.entry_type == "bonus")

    net_payable = round(gross + total_bonuses - total_advances - total_deductions, 2)

    entries_formatted = []
    for e in month_entries:
        entries_formatted.append(
            {
                "id": str(e.id),
                "entry_type": e.entry_type,
                "amount": round(e.amount or 0, 2),
                "date": e.date.isoformat() if e.date else None,
                "note": e.note,
                "created_by": e.created_by,
            }
        )

    # ── Payment history (if payments are stored as PayrollEntry rows) ───────
    payment_history = (
        db.query(PayrollEntry)
        .filter(
            PayrollEntry.worker_id == worker_id,
            PayrollEntry.entry_type == "payment",
        )
        .order_by(PayrollEntry.date.desc())
        .limit(6)
        .all()
    )

    payment_history_formatted = []
    for p in payment_history:
        payment_history_formatted.append(
            {
                "date": p.date.isoformat() if p.date else None,
                "amount": round(p.amount or 0, 2),
                "method": None,
            }
        )

    # ── Earnings history (last 6 months) for charts ─────────────────────────
    earnings_history = []
    for i in range(5, -1, -1):
        yy = today.year
        mm = today.month - i
        while mm <= 0:
            mm += 12
            yy -= 1

        month_records = (
            db.query(AttendanceRecord)
            .filter(
                AttendanceRecord.worker_id == worker_id,
                extract("year", AttendanceRecord.date) == yy,
                extract("month", AttendanceRecord.date) == mm,
                AttendanceRecord.status == "checked_out",
            )
            .all()
        )

        mdays = len(month_records)
        mhours = sum(r.total_hours or 0 for r in month_records)

        if worker.type == "contract":
            mgross = round(math.floor(mhours or 0) * (worker.hourly_rate or 0), 2)
        else:
            mgross = round(mdays * (worker.daily_rate or 0), 2)

        earnings_history.append(
            {
                "month": date(yy, mm, 1).strftime("%b %Y"),
                "gross": mgross,
                "net": mgross,
            }
        )

    return {
        "worker_id": worker_id,
        "month": month_label,
        "present_days": present_days,
        "total_hours": total_hours,
        "rate_label": rate_label,
        "gross": gross,
        "total_advances": round(total_advances, 2),
        "total_deductions": round(total_deductions, 2),
        "total_bonuses": round(total_bonuses, 2),
        "net_payable": net_payable,
        "entries": entries_formatted,
        "payment_history": payment_history_formatted,
        "earnings_history": earnings_history,
    }