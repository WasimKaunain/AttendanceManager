from csv import writer
import pandas as pd
import uuid,os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image


_HEADER_FILL  = "0B1F3A"
_ALT_FILL     = "EEF4FB"


def generate_excel(report_data: dict) -> str:
    file_name = f"/tmp/report_{uuid.uuid4()}.xlsx"

    title    = report_data.get("title", "Report")
    metadata = report_data.get("metadata", {})
    table    = report_data.get("table", {})

    headers = [str(h) for h in table.get("headers", [])]
    rows    = table.get("rows", [])

    total_columns = len(headers)

    with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
        sheet_name = "Report"
        # create sheet safely
        ws = writer.book.create_sheet(title=sheet_name)
        writer.sheets[sheet_name] = ws
        
        # OPTIONAL: remove default sheet if exists
        if "Sheet" in writer.book.sheetnames:
            std = writer.book["Sheet"]
            writer.book.remove(std)

        # ─────────────────────────────────────────────
        # 🟥 TITLE (LEFT ALIGNED FIXED)
        # ─────────────────────────────────────────────
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)

        cell = ws.cell(row=1, column=1)
        cell.value = title
        cell.font = Font(size=18, bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center")  # ✅ FIXED

        ws.row_dimensions[1].height = 30

        # ─────────────────────────────────────────────
        # ⬜ SINGLE BLANK ROW (FIXED)
        # ─────────────────────────────────────────────
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_columns)

        # ─────────────────────────────────────────────
        # 🟦 METADATA (MERGED)
        # ─────────────────────────────────────────────
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=total_columns)

        meta_text = (
            f"Site Name    : {metadata.get('site_name','')}\n"
            f"Date Range   : {metadata.get('date_range','')}\n"
            f"Total Workers: {metadata.get('total_workers','')}\n"
            f"Generated At : {metadata.get('generated_at','')}"
        )

        meta_cell = ws.cell(row=3, column=1)
        meta_cell.value = meta_text
        meta_cell.alignment = Alignment(wrap_text=True, vertical="top")

        ws.row_dimensions[3].height = 85

        # ─────────────────────────────────────────────
        # 🖼️ LOGO (FIXED POSITION)
        # ─────────────────────────────────────────────

        try:
            # backend/app/services/reports → go to backend/
            BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "../../../"))
        
            # go OUT of backend → project root
            PROJECT_ROOT = os.path.abspath(os.path.join(BASE_DIR, ".."))
        
            logo_path = os.path.join(PROJECT_ROOT, "Assets", "AINTSOL_LOGO.png")
        
            print("LOGO PATH:", logo_path)  # debug once
        
            if os.path.exists(logo_path):
                logo = Image(logo_path)
                logo.width = 120
                logo.height = 60
        
                logo_col = get_column_letter(total_columns - 1)
                ws.add_image(logo, f"{logo_col}3")
            else:
                print("Logo NOT found:", logo_path)
        
        except Exception as e:
            print("Logo load failed:", e)

        # ─────────────────────────────────────────────
        # ⬜ EXACTLY ONE BLANK ROW AFTER METADATA
        # ─────────────────────────────────────────────
        ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=total_columns)

        # ─────────────────────────────────────────────
        # 📊 TABLE START (FIXED POSITION)
        # ─────────────────────────────────────────────
        start_row = 5

        df = pd.DataFrame(rows, columns=headers)
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row - 1)

        # ─────────────────────────────────────────────
        # 🎨 STYLING
        # ─────────────────────────────────────────────
        thin = Side(style="thin", color="C8D8EF")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Header
        for col in range(1, total_columns + 1):
            cell = ws.cell(row=start_row, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=_HEADER_FILL)
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        # Data rows
        for row_idx in range(start_row + 1, start_row + 1 + len(rows)):
            fill = _ALT_FILL if row_idx % 2 == 0 else "FFFFFF"

            for col_idx in range(1, total_columns + 1):
                cell = ws.cell(row=row_idx, column=col_idx)

                cell.fill = PatternFill("solid", fgColor=fill)
                cell.border = border
                cell.alignment = Alignment(horizontal="center")

                # 🔴 FIXED RED CROSS
                if str(cell.value).strip() == "✖":
                    cell.value = "✖"
                    cell.font = Font(color="FF0000", bold=True)

        # ─────────────────────────────────────────────
        # 📏 COLUMN WIDTHS
        # ─────────────────────────────────────────────
        ws.column_dimensions["A"].width = 25

        for col in range(2, total_columns - 2):
            ws.column_dimensions[get_column_letter(col)].width = 12

        for col in range(total_columns - 2, total_columns + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # ─────────────────────────────────────────────
        # ❄️ FREEZE HEADER
        # ─────────────────────────────────────────────
        ws.freeze_panes = f"A{start_row + 1}"

    return file_name